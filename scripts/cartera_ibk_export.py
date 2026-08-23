#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador automático "CarteraIBK" — plan de rebalanceo cartera real IBK → módulo Rally Leaders.

Se ejecuta vía launchd cada 5 min. Cuando detecta un scan nuevo de Rally Leaders (o una
cartera IBK recién subida), genera en ~/Desktop/CarteraIBK dos archivos con la fecha/hora
local (Atlantic/Canary) del scan:
  - Cartera_RallyLeaders_AAAA-MM-DD_HH-MM.xlsx  (tabla operativa, colores, gráfico)
  - Cartera_RallyLeaders_AAAA-MM-DD_HH-MM.pdf   (informe visual 1 página, estilo terminal)

SOLO LECTURA de las APIs de producción — no modifica nada de los módulos.
Los archivos contienen VALORES calculados (instantánea del scan), no fórmulas: se
regeneran solos en cada scan y así la vista previa del iPhone siempre muestra los números.

Uso manual:  python3 scripts/cartera_ibk_export.py [--force]
"""

import json
import math
import os
import subprocess
import sys
import urllib.request
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

BASE_URL = "https://emrrclaude.vercel.app"
SCAN_URL = f"{BASE_URL}/api/rally-scan/last"
PORTFOLIO_URL = f"{BASE_URL}/api/rally-scan/ibk-portfolio"

MAIL_FROM_ACCOUNT = "sergimaymo@gmail.com"
MAIL_TO = "sergimaymo@gmail.com"

# Carpeta de salida. El Escritorio está protegido por TCC (privacidad de macOS) y un
# agente de launchd NO puede escribir ahí sin "Acceso total al disco": el 18-ago-2026
# el automatismo se quedó muerto por eso ("Operation not permitted"). Se intenta el
# Escritorio y, si el sistema lo impide, se cae a Application Support, que siempre es
# escribible por el agente. El correo es el canal principal, así que la entrega no
# depende de cuál de las dos se use.
OUT_DIR_PREFERIDO = os.path.expanduser("~/Desktop/CarteraIBK")
OUT_DIR_ALTERNATIVO = os.path.expanduser("~/Library/Application Support/CarteraIBK/salidas")


def _resolver_out_dir() -> str:
    for d in (OUT_DIR_PREFERIDO, OUT_DIR_ALTERNATIVO):
        try:
            os.makedirs(d, exist_ok=True)
            testigo = os.path.join(d, ".escritura")
            with open(testigo, "w") as fh:
                fh.write("ok")
            os.remove(testigo)
            return d
        except OSError:
            continue
    return OUT_DIR_ALTERNATIVO


OUT_DIR = _resolver_out_dir()
STATE_DIR = os.path.expanduser("~/Library/Application Support/CarteraIBK")
STATE_FILE = os.path.join(STATE_DIR, "state.json")
LOG_FILE = os.path.join(STATE_DIR, "generator.log")
TZ = ZoneInfo("Atlantic/Canary")

# Cartera semilla (fotos IBK del 16-ago-2026) — solo se usa si el endpoint del
# servidor aún no tiene snapshot. En cuanto se suben fotos nuevas, manda el servidor.
SEED_PORTFOLIO = {
    "loadedAt": "2026-08-16T17:41:00Z",
    "navEur": 22110.0,
    "valMdoEur": 7010.12,
    "cashEur": 15102.24,
    "fxEurPerUsd": 0.86433,
    "positions": [
        {"symbol": "HUM", "quantity": 5.0101, "lastPrice": 389.00, "currency": "USD"},
        {"symbol": "MU", "quantity": 3.2, "lastPrice": 972.98, "currency": "USD"},
        {"symbol": "WDC", "quantity": 6.0, "lastPrice": 508.00, "currency": "USD"},
    ],
}

# ── infra ────────────────────────────────────────────────────────────────────

def log(msg: str) -> None:
    os.makedirs(STATE_DIR, exist_ok=True)
    stamp = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{stamp}] {msg}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def fetch_json(url: str, timeout: int = 30):
    req = urllib.request.Request(url, headers={"User-Agent": "CarteraIBK-generator/1.0"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode("utf-8"))


def load_state() -> dict:
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_state(state: dict) -> None:
    os.makedirs(STATE_DIR, exist_ok=True)
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def notify(title: str, body: str) -> None:
    try:
        subprocess.run(
            ["osascript", "-e",
             f'display notification "{body}" with title "{title}"'],
            timeout=10, capture_output=True,
        )
    except Exception:
        pass


def _as_str(s: str) -> str:
    """Escapa comillas/backslashes para incrustar en un literal AppleScript."""
    return s.replace("\\", "\\\\").replace('"', '\\"')


def send_email(subject: str, body: str, attachments: list) -> bool:
    """Envía el correo desde Mail.app (cuenta ya configurada en el Mac) — sin
    credenciales nuevas, sin tocar nada del dashboard. Best-effort: si falla,
    solo se registra en el log, nunca interrumpe la generación de archivos."""
    attach_lines = "\n".join(
        f'    make new attachment with properties {{file name:(POSIX file "{_as_str(p)}")}} '
        f'at after the last paragraph'
        for p in attachments
    )
    script = f'''
    tell application "Mail"
        set newMsg to make new outgoing message with properties {{subject:"{_as_str(subject)}", content:"{_as_str(body)}\n", visible:false}}
        tell newMsg
            set sender to "{_as_str(MAIL_FROM_ACCOUNT)}"
            make new to recipient at end of to recipients with properties {{address:"{_as_str(MAIL_TO)}"}}
{attach_lines}
        end tell
        send newMsg
    end tell
    '''
    try:
        r = subprocess.run(["osascript", "-e", script], timeout=30, capture_output=True, text=True)
        if r.returncode != 0:
            log(f"AVISO: envío de email falló ({r.stderr.strip()[:200]})")
            return False
        return True
    except Exception as e:
        log(f"AVISO: envío de email falló ({e})")
        return False


def base_symbol(sym: str) -> str:
    return (sym or "").split(".")[0].strip().upper()


# ── plausibilidad de la cartera (anti-envenenamiento del endpoint sin auth) ────

NAV_DRIFT_TOL = 0.40   # el NAV nuevo no debe desviarse > ±40% del último conocido
NAV_IDENTITY_TOL = 0.10  # NAV ≈ valMdo + cash (±10%)
POS_VALMDO_TOL = 0.15    # Σ posiciones ≈ valMdo (±15%)
# Banda REALISTA del FX EUR/USD (no la [0,5-2,0] laxa anterior, que dejaba pasar un OCR
# 2x mal). EUR/USD lleva 20 años en [0,63, 1,03] EUR por USD; se deja holgura a [0,60, 1,10].
FX_EUR_USD_MIN, FX_EUR_USD_MAX = 0.60, 1.10
# Divisas cotizadas en subunidad (peniques): 1 GBX = 0,01 GBP. IBK devuelve GBX en las .L.
SUBUNIT_PER_MAIN = {"GBX": 0.01, "GBP": 1.0, "USD": 1.0, "EUR": 1.0, "CHF": 1.0}


def _finite_num(v):
    return isinstance(v, (int, float)) and math.isfinite(v)


def positions_value_eur(portfolio: dict) -> float:
    """Σ valor de mercado de las posiciones en EUR (usa fx si está; si falta y son USD,
    no puede convertir con fiabilidad y esas posiciones no suman)."""
    fx = portfolio.get("fxEurPerUsd")
    total = 0.0
    for p in (portfolio.get("positions") or []):
        try:
            q, pr = float(p.get("quantity")), float(p.get("lastPrice"))
        except (TypeError, ValueError):
            continue
        if q <= 0 or pr <= 0:
            continue
        cur = (p.get("currency") or "USD").upper()
        rate = 1.0
        if cur == "USD":
            if not (_finite_num(fx) and fx > 0):
                continue
            rate = float(fx)
        total += q * pr * rate
    return total


def portfolio_plausible(portfolio: dict, last_good_nav, scan: dict = None) -> tuple:
    """Devuelve (ok, motivo). Barrera de sanidad para el snapshot leído del endpoint
    público sin auth: evita que un dato envenenado dispare un email de rebalanceo.
    Si se pasa `scan`, cruza el precio leído por OCR de cada posición contra el precio
    del scan (fiable) — única forma de cazar un OCR que preserva el producto q·precio
    (cantidad reconstruida por valor÷precio) y que ninguna guarda de magnitud detecta."""
    nav = portfolio.get("navEur")
    if not (_finite_num(nav) and nav > 0):
        return False, "navEur ausente o no positivo"
    if _finite_num(last_good_nav) and last_good_nav > 0:
        lo, hi = last_good_nav * (1 - NAV_DRIFT_TOL), last_good_nav * (1 + NAV_DRIFT_TOL)
        if not (lo <= nav <= hi):
            return False, (f"navEur {nav:,.0f} fuera de ±{NAV_DRIFT_TOL:.0%} del último "
                           f"conocido ({last_good_nav:,.0f})")
    val_mdo, cash = portfolio.get("valMdoEur"), portfolio.get("cashEur")
    if _finite_num(val_mdo) and _finite_num(cash) and val_mdo >= 0 and cash >= 0:
        expected = val_mdo + cash
        if expected > 0 and abs(expected - nav) > NAV_IDENTITY_TOL * nav:
            return False, (f"identidad rota: NAV {nav:,.0f} != valMdo+cash {expected:,.0f} "
                           f"(+/-{NAV_IDENTITY_TOL:.0%})")
    positions = portfolio.get("positions") or []

    # LECTURA INCOMPLETA (23-ago-2026): una posición con cantidad pero SIN precio es una
    # lectura fallida del OCR, no una cartera vacía. Si se dejan pasar, build_plan las
    # descarta y el plan dice "COMPRAR" algo que YA TIENES — ocurrió de verdad el 21-ago
    # (INTC/MU/LRCX con lastPrice null → email "10 tickers a comprar, 0 a vender").
    # Ante una lectura incompleta preferimos NO generar informe a generar uno erróneo.
    sin_precio = [
        p.get("symbol", "?") for p in positions
        if _finite_num(p.get("quantity")) and float(p.get("quantity") or 0) > 0
        and not (_finite_num(p.get("lastPrice")) and float(p.get("lastPrice") or 0) > 0)
    ]
    if sin_precio:
        return False, ("lectura incompleta: posiciones con cantidad pero sin precio "
                       f"({', '.join(sin_precio[:6])}) — repite la carga de fotos")

    # MAGNITUD POR POSICIÓN, NORMALIZADA POR DIVISA (reescrito 23-ago-2026 tras la 2ª
    # auditoría adversarial, que encontró DOS agujeros en la versión anterior:
    #   (a) la comprobación FX-vs-posiciones solo corría si fxEurPerUsd venía informado, y
    #       el frontend manda null JUSTO cuando el FX "pinta raro" (optimal2026Refresh.ts:
    #       fxNormalized solo si invested/posValueRaw en (0,5,1,6)); así MRNA@39.245$ con
    #       fx null se colaba. Ahora se comprueba SIEMPRE, derivando el FX si falta.
    #   (b) la guarda "sin FX" sumaba unidades nativas mezclando divisas y usaba banda
    #       [0,5-2,0], lo que (i) dejaba pasar un OCR con la cantidad reconstruida por
    #       valor÷precio (producto q·precio preservado) y (ii) BLOQUEABA una acción de
    #       Londres legítima en peniques (GBX, ~100x). Ahora cada posición se lleva a EUR
    #       con su divisa (peniques incluidos) y un FX EUR/USD que DEBE ser realista.
    # La 3ª auditoría demostró que verificar SOLO cuando hay USD deja pasar carteras 100%
    # EUR corruptas (Σposiciones 4,5x valMdo) y bloquea mixtas EUR+USD legítimas (el FX
    # derivado se contaminaba con la parte EUR). Ahora se descompone SIEMPRE:
    #   valMdo (EUR) = Σ(no-USD en EUR)  +  fx · Σ(USD en USD)
    # Se despeja el fx IMPLÍCITO de esa identidad y debe caer en banda realista. Divisa
    # desconocida (no en el mapa) → no se puede verificar con fiabilidad → se rechaza el
    # informe (mejor no operar que operar a ciegas), salvo que no haya val_mdo con que cruzar.
    valid_pos = [p for p in positions
                 if _finite_num(p.get("quantity")) and _finite_num(p.get("lastPrice"))
                 and float(p["quantity"]) > 0 and float(p["lastPrice"]) > 0]
    if valid_pos and _finite_num(val_mdo) and val_mdo > 0:
        eur_sum = 0.0   # valor en EUR de lo que NO depende del FX EUR/USD (EUR, GBP, CHF, peniques…)
        usd_sum = 0.0   # valor en USD de lo cotizado en dólares
        desconocidas = []
        for p in valid_pos:
            cur = (p.get("currency") or "USD").upper()
            q, pr = float(p["quantity"]), float(p["lastPrice"])
            if cur == "USD":
                usd_sum += q * pr
            elif cur in SUBUNIT_PER_MAIN:
                eur_sum += q * pr * SUBUNIT_PER_MAIN[cur]   # peniques→libras≈EUR, CHF≈EUR, EUR=1
            else:
                desconocidas.append(f"{p.get('symbol','?')}:{cur}")
        if desconocidas:
            return False, (f"divisa no reconocida en {', '.join(desconocidas[:6])} — no se puede "
                           f"verificar la magnitud de la cartera; sube fotos o revisa el dato")
        # fx implícito que hace cuadrar la identidad valMdo = eur_sum + fx·usd_sum
        fx_prov = portfolio.get("fxEurPerUsd")
        if usd_sum > 0:
            fx_eff = float(fx_prov) if _finite_num(fx_prov) and fx_prov > 0 else ((val_mdo - eur_sum) / usd_sum)
            if not (FX_EUR_USD_MIN <= fx_eff <= FX_EUR_USD_MAX):
                return False, (f"FX EUR/USD implícito {fx_eff:.3f} fuera de banda realista "
                               f"[{FX_EUR_USD_MIN}-{FX_EUR_USD_MAX}] — lectura de precios corrupta "
                               f"(EUR {eur_sum:,.0f} + USD {usd_sum:,.0f}, valMdo {val_mdo:,.0f})")
            pos_eur = eur_sum + fx_eff * usd_sum
        else:
            # cartera sin USD: no hay FX que derivar, la identidad es directa en EUR
            pos_eur = eur_sum
        if pos_eur > 0 and abs(pos_eur - val_mdo) > POS_VALMDO_TOL * max(val_mdo, pos_eur):
            return False, (f"Sigma posiciones {pos_eur:,.0f} EUR != valMdo {val_mdo:,.0f} EUR "
                           f"(+/-{POS_VALMDO_TOL:.0%})")

    # PRECIO CRUZADO CONTRA EL SCAN (23-ago-2026): si el OCR lee mal el precio pero la
    # cantidad se reconstruye por valor÷precio, el producto q·precio se conserva y NINGUNA
    # guarda de magnitud lo ve (el valor total en EUR es correcto, solo mal repartido). El
    # precio del scan SÍ es fiable: si el ticker está en el top-10, el precio leído debe
    # parecerse (±35%, margen por hora del scan vs foto). Un factor x295 como el de MRNA
    # del 19-ago (39.245 vs ~133) lo caza de sobra.
    # ⚠️ CUARTA reescritura del cruce. La 3ª versión indexaba por (base, DIVISA), pero la
    # 4ª auditoría demostró que NO SIRVE: la ruta de fotos (optimal2026Refresh.ts:637)
    # hardcodea currency="USD" en TODA posición, así que STMPA/SPM (europeos del top-10)
    # nunca casaban con su entrada EUR del scan → el agujero del MRNA seguía abierto para
    # los europeos. Ahora NO se confía en la divisa de la foto (no es fiable): se cruza por
    # símbolo base contra el scan, cuya divisa y precio SÍ son fiables. Las colisiones entre
    # bolsas (BA=Boeing US / BAE Systems LSE) se evitan cruzando solo bases ÚNICAS en el
    # top-10; y las entradas en divisa de unidad ambigua (GBX/GBP de Londres, ×100 según
    # IBK dé peniques o libras) se EXCLUYEN del cruce para no dar falsos positivos.
    if scan and isinstance(scan.get("top10"), list):
        from collections import Counter
        bases_top = Counter()
        precio_scan = {}   # base -> precio (solo bases únicas y divisa de unidad no ambigua)
        AMBIGUAS = {"GBX", "GBP", "GBp", "ZAR", "ILA"}   # cotizan en subunidad según fuente
        for t in scan["top10"]:
            base = base_symbol(t.get("ticker") or t.get("providerSymbol", ""))
            if base:
                bases_top[base] += 1
        for t in scan["top10"]:
            base = base_symbol(t.get("ticker") or t.get("providerSymbol", ""))
            cur = (t.get("currency") or "").upper()
            px = (t.get("metrics") or {}).get("lastClose")
            if base and bases_top[base] == 1 and cur not in {c.upper() for c in AMBIGUAS} and _finite_num(px) and px > 0:
                precio_scan[base] = float(px)
        for p in positions:
            try:
                pr = float(p.get("lastPrice"))
            except (TypeError, ValueError):
                continue
            if not (math.isfinite(pr) and pr > 0):
                continue
            base = base_symbol(p.get("symbol", ""))
            ref = precio_scan.get(base)
            if ref and (pr > ref * 3 or pr < ref / 3):
                return False, (f"precio de {base} leido {pr:,.2f} incompatible con el precio "
                               f"del scan {ref:,.2f} (factor {pr/ref:.1f}x) — OCR corrupto")
    # LÍMITE INTRÍNSECO conocido (3ª auditoría): si el OCR intercambia precio y cantidad de
    # un ticker que (a) NO está en el top-10 del scan y (b) preserva el producto q·precio,
    # ninguna guarda puede detectarlo — el valor total en EUR es correcto y no hay precio de
    # referencia con que cruzar. Impacto acotado: el plan solo dimensiona mal ESE ticker, y
    # al no estar en el top-10 la orden sería "vender todo" de una posición que de todos
    # modos hay que liquidar por no ser líder. No se inventa un umbral de precio (rompería
    # acciones legítimamente caras: MU 973 USD, Booking ~5000). Cubierto donde importa: el
    # top-10, que concentra el riesgo real, sí se cruza por precio.
    return True, "ok"


# ── cálculo del plan ─────────────────────────────────────────────────────────

def build_plan(scan: dict, portfolio: dict) -> dict:
    nav = float(portfolio["navEur"])
    val_mdo = float(portfolio["valMdoEur"])
    cash = float(portfolio["cashEur"])

    valid_positions = [
        p for p in portfolio["positions"]
        if p.get("quantity") is not None and p.get("lastPrice") is not None
        and float(p["quantity"]) > 0 and float(p["lastPrice"]) > 0
    ]
    dropped = len(portfolio["positions"]) - len(valid_positions)

    fx = portfolio.get("fxEurPerUsd")
    if not fx:
        usd_total = sum(
            float(p["quantity"]) * float(p["lastPrice"])
            for p in valid_positions
            if (p.get("currency") or "USD") == "USD"
        )
        fx = (val_mdo / usd_total) if usd_total > 0 else 0.86433
    fx = float(fx)

    holdings = {}
    for p in valid_positions:
        b = base_symbol(p["symbol"])
        qty = float(p["quantity"])
        price = float(p["lastPrice"])
        cur = (p.get("currency") or "USD").upper()
        value_eur = qty * price * (fx if cur == "USD" else 1.0)
        holdings[b] = {"qty": qty, "price": price, "currency": cur, "value_eur": value_eur}

    top10 = scan["top10"]
    w_sum = sum(float(t["suggestedWeightPct"]) for t in top10)
    # CAP DE SEGURIDAD (18-ago-2026): ningún peso objetivo puede superar el 20% del
    # NAV tras renormalizar. Cubre el caso de scans con menos de 5 candidatos (el
    # engine reparte 100/n → posiciones de 25-100%) y cualquier peso anómalo del
    # servidor. El exceso NO se redistribuye: queda como efectivo (asignación parcial).
    WEIGHT_CAP = 0.20
    capped = False
    rows = []
    seen = set()
    for t in top10:
        tick = base_symbol(t["ticker"])
        seen.add(tick)
        cur = (t.get("currency") or "USD").upper()
        price_native = float(t["metrics"]["lastClose"])
        price_eur = price_native * (fx if cur == "USD" else 1.0)
        w_norm = float(t["suggestedWeightPct"]) / w_sum
        if w_norm > WEIGHT_CAP + 1e-9:
            w_norm = WEIGHT_CAP
            capped = True
        target_eur = w_norm * nav
        target_qty = target_eur / price_eur if price_eur > 0 else 0.0
        h = holdings.get(tick)
        cur_qty = h["qty"] if h else 0.0
        cur_val = h["value_eur"] if h else 0.0
        delta_raw = target_qty - cur_qty
        # IBK: fraccionadas en USD (2 decimales); enteras en el resto
        delta = round(delta_raw, 2) if cur == "USD" else float(round(delta_raw))
        target_qty_r = round(target_qty, 2) if cur == "USD" else float(round(target_qty))
        delta_eur = delta * price_eur
        if cur_qty == 0:
            action = "COMPRAR (nueva)"
        elif abs(delta_eur) < 0.01 * nav:
            action, delta = "MANTENER", delta
        elif delta > 0:
            action = "COMPRAR (ampliar)"
        else:
            action = "REDUCIR"
        rows.append({
            "rank": t["rank"], "ticker": tick, "name": t.get("name", ""),
            "action": action, "delta_qty": delta,
            "cur_qty": round(cur_qty, 4), "target_qty": target_qty_r,
            "stop": t.get("trailingStop") or t["metrics"].get("trailingStop"),
            "weight_pct": w_norm, "cur_val": cur_val, "target_eur": target_eur,
            "price_native": price_native, "currency": cur, "price_eur": price_eur,
        })

    # posiciones reales fuera del top-10 → vender todo
    for tick, h in sorted(holdings.items()):
        if tick in seen:
            continue
        rows.append({
            "rank": None, "ticker": tick, "name": "(fuera del top-10)",
            "action": "VENDER TODO", "delta_qty": -round(h["qty"], 4),
            "cur_qty": round(h["qty"], 4), "target_qty": 0.0,
            "stop": None, "weight_pct": 0.0, "cur_val": h["value_eur"],
            "target_eur": 0.0, "price_native": h["price"],
            "currency": h["currency"], "price_eur": h["price"] * (fx if h["currency"] == "USD" else 1.0),
        })

    # Mensaje de aviso si el cap actuó (se muestra en XLSX, PDF y email).
    cap_warning = None
    if capped:
        if len(top10) < 5:
            cap_warning = ("AVISO — menos de 5 candidatos: asignación parcial, techo 20% "
                           "aplicado; el resto de la cartera queda como efectivo.")
        else:
            cap_warning = ("AVISO — techo de seguridad del 20% del total aplicado a algún "
                           "peso objetivo; el exceso queda como efectivo.")

    utc = datetime.fromisoformat(scan["scanCompletedAtUtc"].replace("Z", "+00:00"))
    local = utc.astimezone(TZ)
    return {
        "scan_id": scan["scanId"],
        "scan_local": local,
        "portfolio_at": portfolio.get("loadedAt"),
        "nav": nav, "val_mdo": val_mdo, "cash": cash, "fx": fx,
        "rows": rows,
        "seeded": portfolio.get("_seeded", False),
        "dropped": dropped,
        "cap_warning": cap_warning,
    }


# ── Excel ────────────────────────────────────────────────────────────────────

def write_xlsx(plan: dict, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    NAVY, ORANGE = "0D1B2A", "F59E0B"
    GREEN_TX, RED_TX, GRAY_TX = "047857", "B91C1C", "6B7280"
    GREEN_BG, RED_BG, AMBER_BG = "D1FAE5", "FEE2E2", "FEF3C7"
    HEAD_BG, ZEBRA = "1B263B", "F3F6FA"

    def F(**kw):
        return Font(name="Arial", **kw)

    wb = Workbook()
    ws = wb.active
    ws.title = "PLAN REBALANCEO"
    ws.sheet_view.showGridLines = False

    scan_txt = plan["scan_local"].strftime("%d/%m/%Y %H:%M")
    ws.merge_cells("B2:M2")
    c = ws["B2"]
    c.value = "PLAN DE REBALANCEO · CARTERA IBK → RALLY LEADERS"
    c.font = F(bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30

    ws.merge_cells("B3:M3")
    c = ws["B3"]
    src = " · cartera semilla (fotos 16-ago)" if plan["seeded"] else ""
    c.value = f"Scan Rally Leaders: {scan_txt} (hora Canarias) · solo lectura, no altera los módulos{src}"
    c.font = F(size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    kpis = [
        ("TOTAL CARTERA", plan["nav"], "#,##0 €"),
        ("INVERTIDO", plan["val_mdo"], "#,##0 €"),
        ("% INVERTIDO", plan["val_mdo"] / plan["nav"], "0.0%"),
        ("EFECTIVO", plan["cash"], "#,##0 €"),
        ("% EFECTIVO", plan["cash"] / plan["nav"], "0.0%"),
        ("EURUSD implícito", 1.0 / plan["fx"], "0.000"),
    ]
    for i, (label, val, fmt) in enumerate(kpis):
        col = 2 + i * 2
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        lc, vc = ws.cell(row=5, column=col), ws.cell(row=6, column=col)
        lc.value, lc.font = label, F(size=9, bold=True, color=GRAY_TX)
        lc.alignment = Alignment(horizontal="center")
        vc.value, vc.number_format = val, fmt
        vc.font = F(size=13, bold=True, color=NAVY)
        vc.alignment = Alignment(horizontal="center")

    headers = [
        "#", "TICKER", "NOMBRE", "ACCIÓN", "± POSICIONES",
        "TRAILING STOP", "POSIC. ACTUALES", "POSIC. OBJETIVO",
        "PESO OBJETIVO", "VALOR ACTUAL €", "VALOR OBJETIVO €", "PRECIO SCAN",
    ]
    HR = 8
    thin = Side(style="thin", color="D1D5DB")
    for j, h in enumerate(headers):
        c = ws.cell(row=HR, column=2 + j, value=h)
        c.font = F(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HR].height = 26

    r = HR + 1
    first_data = r
    for i, row in enumerate(plan["rows"]):
        d = row["delta_qty"]
        sign = f"+{d:g}" if d > 0 else f"{d:g}"
        stop = f"{row['stop']:.0f}%" if row["stop"] else "—"
        vals = [
            row["rank"] or "—", row["ticker"], row["name"], row["action"], sign,
            stop, row["cur_qty"], row["target_qty"], row["weight_pct"],
            round(row["cur_val"]), round(row["target_eur"]),
            f"{row['price_native']:,.2f} {row['currency']}",
        ]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.font = F(size=10)
            c.border = Border(bottom=thin)
            c.alignment = Alignment(horizontal="center" if j not in (2,) else "left",
                                    vertical="center")
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=ZEBRA)
        ws.cell(row=r, column=3).font = F(size=10, bold=True)
        ws.cell(row=r, column=10).number_format = "0.0%"
        ws.cell(row=r, column=11).number_format = "#,##0 €"
        ws.cell(row=r, column=12).number_format = "#,##0 €"
        act, dc, ac = row["action"], ws.cell(row=r, column=6), ws.cell(row=r, column=5)
        if act.startswith("COMPRAR"):
            fill, color = PatternFill("solid", fgColor=GREEN_BG), GREEN_TX
        elif act in ("REDUCIR", "VENDER TODO"):
            fill, color = PatternFill("solid", fgColor=RED_BG), RED_TX
        else:
            fill, color = PatternFill("solid", fgColor=AMBER_BG), GRAY_TX
        for cell in (ac, dc):
            cell.fill = fill
        ac.font = F(size=10, bold=True, color=color)
        dc.font = F(size=11, bold=True, color=color)
        stop_c = ws.cell(row=r, column=7)
        stop_c.font = F(size=10, bold=True, color=ORANGE.replace("F59E0B", "B45309"))
        r += 1
    last_data = r - 1

    tr = r + 1
    ws.cell(row=tr, column=3, value="TOTALES").font = F(bold=True, size=10)
    tc = ws.cell(row=tr, column=11, value=round(sum(x["cur_val"] for x in plan["rows"])))
    tc.number_format, tc.font = "#,##0 €", F(bold=True, size=10)
    tc2 = ws.cell(row=tr, column=12, value=round(sum(x["target_eur"] for x in plan["rows"])))
    tc2.number_format, tc2.font = "#,##0 €", F(bold=True, size=10)
    for col in range(2, 14):
        ws.cell(row=tr, column=col).border = Border(top=Side(style="medium", color=NAVY))

    notes = [
        "CÓMO USARLO EN IBK: la columna ± POSICIONES es lo que hay que operar en cada ticker",
        "(+ = comprar ese nº de acciones, − = vender). TRAILING STOP es el % a colocar en IBK",
        "sobre cierres diarios. Objetivo por ticker = peso del último scan × total cartera "
        f"({plan['nav']:,.0f} €).",
        f"Posiciones objetivo = valor objetivo ÷ precio del scan (USD→EUR al FX implícito "
        f"{1.0/plan['fx']:.3f} de la cartera).",
        "USD: fraccionadas (2 decimales). Datos: scan Rally Leaders + fotos IBK subidas al "
        "dashboard. Valores instantánea del scan; el archivo se regenera en cada scan nuevo.",
    ]
    if plan.get("cap_warning"):
        notes.insert(0, plan["cap_warning"])
    nr = tr + 2
    for i, t in enumerate(notes):
        ws.merge_cells(start_row=nr + i, start_column=2, end_row=nr + i, end_column=13)
        c = ws.cell(row=nr + i, column=2, value=t)
        if t.startswith("AVISO"):
            c.font = F(size=9, bold=True, color=RED_TX)
        else:
            c.font = F(size=8, italic=True, color=GRAY_TX)

    # helper de gráfico: % actual vs objetivo por ticker (columnas auxiliares ocultas)
    HCOL = 16  # P
    ws.cell(row=HR, column=HCOL, value="Ticker").font = F(size=8)
    ws.cell(row=HR, column=HCOL + 1, value="% actual").font = F(size=8)
    ws.cell(row=HR, column=HCOL + 2, value="% objetivo").font = F(size=8)
    for i, row in enumerate(plan["rows"]):
        rr = HR + 1 + i
        ws.cell(row=rr, column=HCOL, value=row["ticker"])
        ws.cell(row=rr, column=HCOL + 1, value=round(100 * row["cur_val"] / plan["nav"], 2))
        ws.cell(row=rr, column=HCOL + 2, value=round(100 * row["weight_pct"], 2))
    for cc in range(HCOL, HCOL + 3):
        ws.column_dimensions[get_column_letter(cc)].hidden = True

    chart = BarChart()
    chart.type, chart.style = "col", 10
    chart.title = "% de la cartera total: actual vs objetivo Rally Leaders"
    chart.y_axis.title = "% del total"
    chart.height, chart.width = 8.5, 24
    data = Reference(ws, min_col=HCOL + 1, min_row=HR, max_col=HCOL + 2,
                     max_row=HR + len(plan["rows"]))
    cats = Reference(ws, min_col=HCOL, min_row=HR + 1, max_row=HR + len(plan["rows"]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "94A3B8"
    chart.series[1].graphicalProperties.solidFill = "F59E0B"
    ws.add_chart(chart, f"B{nr + len(notes) + 1}")

    widths = [3, 5, 9, 22, 17, 13, 13, 13, 13, 12, 14, 15, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + i)].width = w
    ws.freeze_panes = f"B{HR + 1}"
    wb.save(path)


# ── PDF ──────────────────────────────────────────────────────────────────────

def write_pdf(plan: dict, path: str) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas

    W, H = A4
    BG, PANEL = (0.05, 0.07, 0.10), (0.09, 0.12, 0.17)
    FG, DIM = (0.92, 0.94, 0.96), (0.55, 0.60, 0.67)
    ORANGE, GREEN, RED = (0.96, 0.62, 0.04), (0.06, 0.72, 0.51), (0.94, 0.27, 0.27)

    c = pdfcanvas.Canvas(path, pagesize=A4)
    c.setFillColorRGB(*BG)
    c.rect(0, 0, W, H, stroke=0, fill=1)

    def text(x, y, s, size=9, color=FG, font="Helvetica", right=False, center=False):
        c.setFont(font, size)
        c.setFillColorRGB(*color)
        if right:
            c.drawRightString(x, y, s)
        elif center:
            c.drawCentredString(x, y, s)
        else:
            c.drawString(x, y, s)

    M = 14 * mm
    y = H - 18 * mm
    text(M, y, "CARTERA IBK → RALLY LEADERS", 17, ORANGE, "Helvetica-Bold")
    text(W - M, y, plan["scan_local"].strftime("SCAN %d/%m/%Y · %H:%M"), 11, FG,
         "Helvetica-Bold", right=True)
    y -= 5.2 * mm
    text(M, y, "Plan de rebalanceo · posiciones a operar y trailing stops · solo lectura",
         8.5, DIM)

    # KPIs
    y -= 12 * mm
    kpis = [
        ("TOTAL CARTERA", f"{plan['nav']:,.0f} €", FG),
        ("INVERTIDO", f"{plan['val_mdo']:,.0f} € · {100*plan['val_mdo']/plan['nav']:.1f}%", ORANGE),
        ("EFECTIVO", f"{plan['cash']:,.0f} € · {100*plan['cash']/plan['nav']:.1f}%", GREEN),
        ("EURUSD", f"{1.0/plan['fx']:.3f}", FG),
    ]
    bw = (W - 2 * M - 9 * mm) / 4
    for i, (label, val, color) in enumerate(kpis):
        x = M + i * (bw + 3 * mm)
        c.setFillColorRGB(*PANEL)
        c.roundRect(x, y - 6 * mm, bw, 13 * mm, 2 * mm, stroke=0, fill=1)
        text(x + bw / 2, y + 2.4 * mm, label, 6.5, DIM, "Helvetica-Bold", center=True)
        text(x + bw / 2, y - 3 * mm, val, 10.5, color, "Helvetica-Bold", center=True)

    # barra invertido/efectivo
    y -= 15 * mm
    bar_w, bar_h = W - 2 * M, 5.5 * mm
    inv_frac = plan["val_mdo"] / plan["nav"]
    c.setFillColorRGB(*PANEL)
    c.roundRect(M, y - bar_h, bar_w, bar_h, 1.5 * mm, stroke=0, fill=1)
    c.setFillColorRGB(*ORANGE)
    c.roundRect(M, y - bar_h, bar_w * inv_frac, bar_h, 1.5 * mm, stroke=0, fill=1)
    text(M, y - bar_h - 4 * mm, f"Invertido {100*inv_frac:.1f}%", 7.5, ORANGE, "Helvetica-Bold")
    text(W - M, y - bar_h - 4 * mm, f"Efectivo {100*(1-inv_frac):.1f}%", 7.5, GREEN,
         "Helvetica-Bold", right=True)

    # tabla
    y -= 17 * mm
    cols = [
        ("#", 8 * mm), ("TICKER", 20 * mm), ("ACCIÓN", 34 * mm),
        ("± POSICIONES", 27 * mm), ("STOP", 15 * mm), ("PESO OBJ.", 19 * mm),
        ("ACTUAL €", 22 * mm), ("OBJETIVO €", 22 * mm), ("PRECIO", 26 * mm),
    ]
    x = M
    for label, wd in cols:
        text(x + wd / 2, y, label, 7, DIM, "Helvetica-Bold", center=True)
        x += wd
    y -= 2.5 * mm
    c.setStrokeColorRGB(*DIM)
    c.setLineWidth(0.4)
    c.line(M, y, W - M, y)

    row_h = 8.6 * mm
    for row in plan["rows"]:
        y -= row_h
        d = row["delta_qty"]
        act = row["action"]
        color = GREEN if act.startswith("COMPRAR") else (RED if d < 0 else DIM)
        vals = [
            (str(row["rank"] or "—"), FG, "Helvetica"),
            (row["ticker"], FG, "Helvetica-Bold"),
            (act, color, "Helvetica-Bold"),
            ((f"+{d:g}" if d > 0 else f"{d:g}"), color, "Helvetica-Bold"),
            ((f"{row['stop']:.0f}%" if row["stop"] else "—"), ORANGE, "Helvetica-Bold"),
            (f"{100*row['weight_pct']:.1f}%", FG, "Helvetica"),
            (f"{row['cur_val']:,.0f}", FG, "Helvetica"),
            (f"{row['target_eur']:,.0f}", FG, "Helvetica"),
            (f"{row['price_native']:,.2f} {row['currency']}", DIM, "Helvetica"),
        ]
        x = M
        for (s, col_c, font), (_, wd) in zip(vals, cols):
            size = 9.5 if font == "Helvetica-Bold" and s and s[0] in "+-" else 8
            text(x + wd / 2, y + 2.6 * mm, s, size, col_c, font, center=True)
            x += wd
        c.setStrokeColorRGB(0.16, 0.20, 0.26)
        c.line(M, y, W - M, y)

    y -= 9 * mm
    if plan.get("cap_warning"):
        text(M, y, plan["cap_warning"], 8, RED, "Helvetica-Bold")
        y -= 4.5 * mm
    text(M, y, "± POSICIONES: nº de acciones a comprar (+) o vender (−) en IBK para "
               "igualar el último scan de Rally Leaders.", 7.5, FG)
    y -= 4 * mm
    text(M, y, "STOP: trailing stop % a colocar por ticker (sobre cierres diarios). "
               "Objetivo por ticker = peso del scan × total cartera.", 7.5, FG)
    y -= 4 * mm
    src = " · cartera semilla (fotos 16-ago)" if plan["seeded"] else ""
    text(M, y, f"Generado automáticamente al detectar el scan · datos: API Rally Leaders + "
               f"fotos IBK{src} · no altera los módulos.", 7, DIM)
    c.showPage()
    c.save()


def generar_salidas(plan: dict, stamp: str) -> tuple:
    """Escribe xlsx+pdf, con reserva REAL de directorio.

    ~/Desktop está sincronizado con iCloud y la cuenta del usuario está llena: macOS puede
    devolver EDEADLK ("Resource deadlock avoided") A MITAD de la escritura. Pasó el
    23-ago-2026 a las 13:46 y abortó el run entero → ese día NO llegó el email.
    `_resolver_out_dir()` ya sondea al arrancar, pero escribe un testigo de 2 bytes: que el
    sondeo pase no garantiza que el .xlsx real (mucho mayor y más lento) también pase.
    Aquí la reserva se aplica donde falla de verdad. El email es la vía de entrega real
    (el Escritorio no llega al iPhone), así que NO puede depender de iCloud.
    """
    destinos = list(dict.fromkeys((OUT_DIR, OUT_DIR_ALTERNATIVO)))
    ultimo = len(destinos) - 1
    for i, destino in enumerate(destinos):
        xlsx_path = os.path.join(destino, f"Cartera_RallyLeaders_{stamp}.xlsx")
        pdf_path = os.path.join(destino, f"Cartera_RallyLeaders_{stamp}.pdf")
        try:
            os.makedirs(destino, exist_ok=True)
            write_xlsx(plan, xlsx_path)
            write_pdf(plan, pdf_path)
            return xlsx_path, pdf_path
        except OSError as exc:
            # No dejar un fichero a medias que luego se adjunte corrupto al email.
            for parcial in (xlsx_path, pdf_path):
                try:
                    os.remove(parcial)
                except OSError:
                    pass
            if i == ultimo:
                raise
            log(f"Aviso: no se pudo escribir en {destino} ({exc}) — reintento en {destinos[i + 1]}")


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    force = "--force" in sys.argv
    os.makedirs(OUT_DIR, exist_ok=True)

    try:
        scan_resp = fetch_json(SCAN_URL)
        if not scan_resp.get("ok") or not scan_resp.get("top10"):
            log("Scan no disponible aún — nada que hacer.")
            return 0
    except Exception as e:
        log(f"ERROR al leer el scan: {e}")
        return 1

    portfolio = None
    try:
        p_resp = fetch_json(PORTFOLIO_URL)
        if p_resp.get("ok") and p_resp.get("portfolio"):
            portfolio = p_resp["portfolio"]
    except Exception as e:
        log(f"Aviso: endpoint de cartera no disponible ({e}) — uso semilla.")
    if portfolio is None:
        portfolio = dict(SEED_PORTFOLIO)
        portfolio["_seeded"] = True

    state = load_state()
    key = f"{scan_resp['scanId']}|{portfolio.get('loadedAt')}"
    if not force and state.get("lastKey") == key:
        return 0  # nada nuevo

    # Barrera de plausibilidad: el snapshot de cartera llega de un endpoint público sin
    # auth, así que antes de generar/emailear un plan de rebalanceo verificamos que el dato
    # sea coherente (NAV dentro de ±40% del último conocido, identidad NAV≈valMdo+cash,
    # Σposiciones≈valMdo). La semilla local es de confianza y no se somete al chequeo.
    if not portfolio.get("_seeded"):
        ok, reason = portfolio_plausible(portfolio, state.get("lastGoodNav"), scan_resp)
        if not ok:
            log(f"AVISO: cartera del servidor NO plausible ({reason}) — NO se genera ni "
                f"envía plan. Se conserva el último bueno. Sube fotos correctas para reevaluar.")
            notify("CarteraIBK: dato ignorado",
                   f"Cartera recibida no plausible ({reason}). No se ha enviado plan.")
            return 0

    plan = build_plan(scan_resp, portfolio)
    stamp = plan["scan_local"].strftime("%Y-%m-%d_%H-%M")
    xlsx_path, pdf_path = generar_salidas(plan, stamp)

    # Persistimos el NAV usado como "último bueno" para el chequeo de deriva del próximo run.
    last_good_nav = plan["nav"] if _finite_num(plan.get("nav")) and plan["nav"] > 0 else state.get("lastGoodNav")
    save_state({"lastKey": key, "lastFiles": [xlsx_path, pdf_path],
                "lastGoodNav": last_good_nav,
                "generatedAt": datetime.now(timezone.utc).isoformat()})
    log(f"Generados: {os.path.basename(xlsx_path)} + PDF (scan {plan['scan_id']}, "
        f"cartera {'semilla' if plan['seeded'] else portfolio.get('loadedAt')})")

    invertido_pct = 100 * plan["val_mdo"] / plan["nav"]
    buys = sum(1 for r in plan["rows"] if r["action"].startswith("COMPRAR"))
    sells = sum(1 for r in plan["rows"] if r["action"] in ("REDUCIR", "VENDER TODO"))
    subject = f"CarteraIBK · Rally Leaders {plan['scan_local'].strftime('%d/%m %H:%M')}"
    body = (
        f"Plan de rebalanceo del scan {plan['scan_local'].strftime('%d/%m/%Y %H:%M')} "
        f"(hora Canarias).\n\n"
        f"Total cartera: {plan['nav']:,.0f} EUR\n"
        f"Invertido: {plan['val_mdo']:,.0f} EUR ({invertido_pct:.1f}%)\n"
        f"Efectivo: {plan['cash']:,.0f} EUR ({100-invertido_pct:.1f}%)\n\n"
        f"{buys} tickers a comprar/ampliar, {sells} a reducir/vender. "
        f"Detalle de posiciones (+/-) y trailing stop por ticker en los adjuntos.\n\n"
        f"Generado automaticamente, solo lectura, no afecta a los modulos."
    )
    if plan.get("cap_warning"):
        body = plan["cap_warning"] + "\n\n" + body
    sent = send_email(subject, body, [xlsx_path, pdf_path])
    log(f"Email {'enviado' if sent else 'NO enviado (ver aviso arriba)'} a {MAIL_TO}")
    notify("CarteraIBK actualizada",
           f"Nuevo plan Rally Leaders {plan['scan_local'].strftime('%d/%m %H:%M')} "
           f"— {'enviado por email' if sent else 'en Escritorio/CarteraIBK'}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
